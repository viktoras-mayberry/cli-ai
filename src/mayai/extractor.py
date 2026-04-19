"""Extract text from various file formats for reading and summarization."""

from __future__ import annotations

import csv
import io
import json
import os
from pathlib import Path

MAX_CHARS = 50_000


def extract_text(filepath: str, max_chars: int = MAX_CHARS) -> tuple[str, str]:
    """Extract text from a file. Returns (text_content, format_description).

    Supports: .txt, .md, .csv, .json, .yaml/.yml, .toml, .pdf, .docx, .xlsx,
    and falls back to raw text read for unknown extensions.

    Libraries like pdfplumber, python-docx, openpyxl are optional;
    if not installed, a helpful error message is returned.
    """
    path = Path(filepath)
    if not path.exists():
        return "", f"File not found: {filepath}"

    ext = path.suffix.lower()
    size = path.stat().st_size

    try:
        if ext in (".txt", ".md", ".log", ".rst", ".ini", ".cfg", ".sh", ".bat",
                    ".py", ".js", ".ts", ".html", ".css", ".xml", ".sql", ".r"):
            return _read_text(path, max_chars), f"Plain text ({ext})"

        if ext == ".csv":
            return _read_csv(path, max_chars), "CSV spreadsheet"

        if ext == ".json":
            return _read_json(path, max_chars), "JSON data"

        if ext in (".yaml", ".yml"):
            return _read_yaml(path, max_chars), "YAML data"

        if ext == ".toml":
            return _read_text(path, max_chars), "TOML config"

        if ext == ".pdf":
            return _read_pdf(path, max_chars), "PDF document"

        if ext in (".docx", ".doc"):
            return _read_docx(path, max_chars), "Word document"

        if ext in (".xlsx", ".xls"):
            return _read_xlsx(path, max_chars), "Excel spreadsheet"

        return _read_text(path, max_chars), f"File ({ext or 'no extension'})"

    except Exception as exc:
        return "", f"Failed to extract text: {exc}"


def _read_text(path: Path, max_chars: int) -> str:
    encodings = ["utf-8", "latin-1", "cp1252"]
    for enc in encodings:
        try:
            text = path.read_text(encoding=enc)
            return text[:max_chars]
        except (UnicodeDecodeError, ValueError):
            continue
    return "(Could not decode file with common encodings)"


def _read_csv(path: Path, max_chars: int) -> str:
    lines: list[str] = []
    total = 0
    with open(path, newline="", encoding="utf-8", errors="replace") as f:
        reader = csv.reader(f)
        for row in reader:
            line = " | ".join(row)
            if total + len(line) > max_chars:
                lines.append("... (truncated)")
                break
            lines.append(line)
            total += len(line) + 1
    return "\n".join(lines)


def _read_json(path: Path, max_chars: int) -> str:
    raw = path.read_text(encoding="utf-8", errors="replace")
    try:
        data = json.loads(raw)
        formatted = json.dumps(data, indent=2, ensure_ascii=False)
        return formatted[:max_chars]
    except json.JSONDecodeError:
        return raw[:max_chars]


def _read_yaml(path: Path, max_chars: int) -> str:
    return _read_text(path, max_chars)


def _read_pdf(path: Path, max_chars: int) -> str:
    try:
        import pdfplumber
    except ImportError:
        return (
            "(PDF support requires pdfplumber. "
            "Install it with: pip install pdfplumber)"
        )

    text_parts: list[str] = []
    total = 0
    with pdfplumber.open(str(path)) as pdf:
        for page in pdf.pages:
            page_text = page.extract_text() or ""
            if total + len(page_text) > max_chars:
                text_parts.append(page_text[:max_chars - total])
                text_parts.append("... (truncated)")
                break
            text_parts.append(page_text)
            total += len(page_text)
    return "\n\n".join(text_parts)


def _read_docx(path: Path, max_chars: int) -> str:
    try:
        from docx import Document
    except ImportError:
        return (
            "(Word document support requires python-docx. "
            "Install it with: pip install python-docx)"
        )

    doc = Document(str(path))
    text_parts: list[str] = []
    total = 0
    for para in doc.paragraphs:
        text = para.text
        if total + len(text) > max_chars:
            text_parts.append(text[:max_chars - total])
            text_parts.append("... (truncated)")
            break
        text_parts.append(text)
        total += len(text)
    return "\n".join(text_parts)


def _read_xlsx(path: Path, max_chars: int) -> str:
    try:
        from openpyxl import load_workbook
    except ImportError:
        return (
            "(Excel support requires openpyxl. "
            "Install it with: pip install openpyxl)"
        )

    wb = load_workbook(str(path), read_only=True, data_only=True)
    lines: list[str] = []
    total = 0
    for sheet_name in wb.sheetnames:
        lines.append(f"--- Sheet: {sheet_name} ---")
        ws = wb[sheet_name]
        for row in ws.iter_rows(values_only=True):
            line = " | ".join(str(cell) if cell is not None else "" for cell in row)
            if total + len(line) > max_chars:
                lines.append("... (truncated)")
                wb.close()
                return "\n".join(lines)
            lines.append(line)
            total += len(line) + 1
    wb.close()
    return "\n".join(lines)
